import pandas as panda
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def exportar_datos_a_excel(origen_datos, nombre_salida="Reporte_ventas.xlsx"):



    print(f"Procesando datos para {nombre_salida}")

    df = None
    #!Sacamos los datos de la db mediante una query
    if isinstance(origen_datos,str) and origen_datos.endswith((".db")):
        if not os.path.exists(origen_datos):
            print(f"Error no existe la base de datos en {origen_datos}")
            return
        try:
            conn = sqlite3.connect(origen_datos)
            query = """
                    SELECT 
                        v.id_venta,
                        v.fecha,
                        v.cliente,
                        v.telefono,
                        v.referencia,
                        -- A partir de aquí son datos del DETALLE
                        d.producto,
                        d.cantidad,
                        d.precio as precio_unitario,
                        (d.cantidad * d.precio) as subtotal,
                        -- Total global de la venta (para referencia)
                        v.total as total
                    FROM ventas v
                    LEFT JOIN detalles d ON v.id_venta = d.id_venta
                    ORDER BY v.id_venta DESC
                    """
            df = panda.read_sql_query(query, conn)
            conn.close()
            print(f"Datos extraidos correctamente de la base de datos")
        except Exception as e:
            print(f"Error leyendo la sql: {e}")
            return
        #! Se para aca si no encuentra la db
    else:
        print(f"Error el origen de los datos no es un archivo valido")
        return
    #!Calcula el total final segun los subtotales individuales
    gran_total = df['subtotal'].sum()
    
    #!Guardamos en el excel los datos
    try:
        df.to_excel(nombre_salida, index=False, sheet_name="Datos_Reales", engine="openpyxl")
    except Exception as e:
        print(f"Error escribiendo el excel: {e}")
    #! 3. APLICAR FORMATO "LEGIBLE" (Auto-ancho y cabeceras)
    try:
        wb = load_workbook(nombre_salida)
        ws = wb.active

        # Estilo para la cabecera (Negrita, fondo gris claro, borde)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))


        FORMATO_MONEDA = '"$"#,##0.00_-'

        COLUMNAS_DINERO = ["precio_unitario","subtotal","total"]

        # 1. Primero detectamos dónde terminan los datos actuales
        ultima_fila_datos = ws.max_row
        ultima_columna_letra = get_column_letter(ws.max_column)

        # 2. Aplicamos el filtro HASTA esa fila (sin incluir el total que pondremos después)
        rango_filtro = f"A1:{ultima_columna_letra}{ultima_fila_datos}"
        ws.auto_filter.ref = rango_filtro
        
        # 3. Ahora sí, definimos que la fila del "Total General" será la siguiente
        fila_final = ultima_fila_datos + 1

        #! Se agrega la fila de suma total
        # Encontramos la primera fila vacía después de los datos
        fila_final = ws.max_row + 1
        
        # Escribimos el título "TOTAL GENERAL" (ajusta la letra de columna según tu gusto)
        # Suponiendo que la columna 'H' es subtotal_linea (index 8)
        # Letra G es Precio, Letra H es Subtotal.
        
        celda_titulo = ws[f"I{fila_final}"]
        celda_titulo.value = "TOTAL GENERAL:"
        celda_titulo.font = Font(bold=True, size=12)
        celda_titulo.alignment = Alignment(horizontal='right')

        celda_valor = ws[f"J{fila_final}"]
        celda_valor.value = gran_total
        celda_valor.font = Font(bold=True, size=12, color="0000FF") # Azul y Negrita
        celda_valor.number_format = '"$"#,##0.00_-' # Formato Moneda
        celda_valor.border = Border(top=Side(style='double')) 

        for column in ws.columns:
            header_cell = column[0]
            column_letter = column[0].column_letter
            nombre_columna = header_cell.value

            #?estilos
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.border = thin_border

        # Auto-ajustar ancho de columnas
            max_length = 0            
            for cell in column:
                try:
                    # Aplicar formato moneda a los datos (filas intermedias)
                    # Ojo: No sobreescribir el formato del total general (fila final)
                    if (cell.row > 1 and cell.row < fila_final) and nombre_columna in COLUMNAS_DINERO:
                        cell.number_format = FORMATO_MONEDA
                    
                    if cell.value:
                        longitud_dato = len(str(cell.value))
                        if longitud_dato > max_length:
                            max_length = longitud_dato
                except:
                    pass
            
            # Ajustamos el ancho (con un poco de margen extra)
            adjusted_width = (max_length + 2) * 1.2
            if nombre_columna in COLUMNAS_DINERO:
                adjusted_width +=4
            ws.column_dimensions[column_letter].width = max(adjusted_width,10)


        #!Guardamos el archivo
        wb.save(nombre_salida)
        print(f"✨ Excel generado y formateado exitosamente: {nombre_salida}")

    except Exception as e:
        print(f"⚠️ El Excel se creó, pero falló el formato visual: {e}")